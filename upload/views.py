from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse

from .forms import UploadForm
from django.conf import settings
import os

import openpyxl, re

def index(request):
    """The home page which generates the donor list html page"""
    if request.method != 'POST':
        form = UploadForm()
    else:
        form = UploadForm(request.POST or None, request.FILES or None)
        if form.is_valid():
            form.save()
            file = request.FILES['fileobj'].name
            file_corrected = file.replace(" ", "_")
            path = os.path.join(settings.MEDIA_ROOT, file_corrected)
            wb = openpyxl.load_workbook(path)
            sheet = wb.get_sheet_by_name('Sheet1')
            text_file = open('upload/templates/upload/donor_list.html', 'w')

            html1 = "{% extends 'upload/base.html' %}" + "\n" + "{% block header %}" + "\n" + "  <h1>Donor List</h1>" + "\n" + "{% endblock header %}" + "\n" + "{% block content %}" + "\n"
            html2 = "{% endblock content %}"
            text_file.write(html1)

            for rowNum in range(1, sheet.max_row + 1):
                firstName = str(sheet.cell(row=rowNum, column=1).value)
                if firstName == "None":
                    firstName = "\n"
                lastName = str(sheet.cell(row=rowNum, column=2).value)
                addNum = re.compile(r'\d(\d)*')
                addressNumber1 = addNum.search(str(sheet.cell(row=rowNum, column=3).value))
                if addressNumber1 is None:
                    addressNumber = ""
                if addressNumber1 is not None:
                    addressNumber = addressNumber1.group(0)
                donate = str(sheet.cell(row=rowNum, column=4).value)
                donate = "$" + donate
                if donate == "$None":
                    donate = ""
                date = str(sheet.cell(row=rowNum, column=5).value)
                year = date[2:4]
                if year == "ne":
                    year = ""
                if firstName == "\n" and lastName != "None":
                    firstName = str(sheet.cell(row=rowNum, column=2).value)
                    lastName = str(sheet.cell(row=rowNum, column=3).value)
                    addressNumber1 = addNum.search(str(sheet.cell(row=rowNum, column=4).value))
                    addressNumber = addressNumber1.group(0)
                    donate = str(sheet.cell(row=rowNum, column=5).value)
                    donate = "$" + donate
                    date = str(sheet.cell(row=rowNum, column=6).value)
                    year = date[2:4]
                if firstName is "_" or lastName is "Anonymous":
                    text_file.write("""  <p>{} (Mr./Ms.) {} {} {}</p>""".format(addressNumber, lastName, donate, year) + '\n')

                else:
                    text_file.write("""  <p>{} {} {} {}</p>""".format(addressNumber, firstName, donate, year) + '\n')

            text_file.write(html2)
            text_file.close()
            os.remove(path)
            return donor_list(request)
    context = {'form': form}
    return render(request, 'upload/index.html', context)

def instructions(request):
    """The how-to for file uploading"""
    return render(request, 'upload/instructions.html')

def donor_list(request):
    """Webpage with the donor list after upload"""
    return render(request, 'upload/donor_list.html')
